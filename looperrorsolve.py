import frappe
from frappe import exceptions as frappe_exceptions
from frappe.utils import get_site_path
from frappe.utils.file_manager import save_file
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import datetime
import re
import json
import csv
import io
import base64
import tempfile
import os
import traceback
import time

# ================= CONSTANTS & STYLES =================

MAX_FILE_SIZE_MB = 20
VALIDATION_TIMEOUT = 120
SHEET_TIMEOUT = 60
LINK_CACHE_LIMIT = 5000  # Adjust this if needed

ERROR_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
ERROR_TEXT_COLOR = Font(color="9C0006")
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

ERROR_CODE_LABELS = {
    "LINK_NOT_FOUND": "Data Not Found",
    "INVALID_YEAR_RANGE": "Invalid Year Range",
    "REQUIRED_FIELD_EMPTY": "Required Field Empty",
    "INVALID_INT": "Must Be a Number",
    "DUPLICATE_ROW": "Duplicate Row",
    "EMPTY_ROW": "Empty Row",
    "INVALID_FLOAT": "Must Be a Decimal Number",
    "INVALID_DATE": "Invalid Date",
    "INVALID_DATETIME": "Invalid Date/Time",
    "INVALID_YEAR": "Invalid Year Format",
    "INVALID_TEXT": "Invalid Text Format",
    "INVALID_SELECT": "Invalid Selection",
    "INVALID_LINK": "Invalid Link",
    "DUPLICATE_PRIMARY_KEY": "Duplicate ID",
    "DUPLICATE_UNIQUE": "Duplicate Value",
    "NO_FILE": "No File Uploaded",
    "INVALID_FILE_TYPE": "Invalid File Type (Use .xlsx)",
    "FILE_TOO_LARGE": "File Too Large",
    "PASSWORD_PROTECTED": "Password Protected File",
    "UNREADABLE_FILE": "Corrupted File",
    "SHEET_NOT_FOUND": "Worksheet Not Found",
    "NO_HEADER": "Header Row Missing",
    "EMPTY_HEADERS": "Empty Header Row",
    "DUPLICATE_HEADERS": "Duplicate Headers",
    "NO_DATA_ROWS": "No Data Rows",
    "MISSING_REQUIRED_COLUMNS": "Missing Required Columns",
    "DOCTYPE_NOT_FOUND": "DocType Not Found",
    "NO_PERMISSION": "No Permission",
    "DOCTYPE_ERROR": "DocType Error",
    "LINK_CONFIG_ERROR": "Configuration Error",
    "INVALID_LINK_TYPE": "Invalid Link Type",
    "LINK_DOCTYPE_NOT_FOUND": "Target Type Not Found",
    "LINK_PERMISSION_ERROR": "Permission Denied",
    "LINK_VALIDATION_ERROR": "Link Validation Error",
    "TIMEOUT_ERROR": "Processing Timeout",
}

_link_cache = {}
_link_cache_sizes = {}


class ValidationTimeout(Exception):
    pass


# ================= MAIN FUNCTION =================

@frappe.whitelist(allow_guest=True)
def validate_and_add_error_columns(update="0", skip_link_validation="0"):
    """
    WITH LINK VALIDATION ENABLED - Detailed logging version
    """
    skip_links = skip_link_validation == "1"
    start_time = time.time()
    
    # Log configuration
    log_msg = f"""
{'='*60}
VALIDATION STARTED
{'='*60}
Skip Link Validation: {skip_links}
Link Cache Limit: {LINK_CACHE_LIMIT:,} records
Sheet Timeout: {SHEET_TIMEOUT}s
Total Timeout: {VALIDATION_TIMEOUT}s
{'='*60}
"""
    print(log_msg)
    frappe.log_error(log_msg, "Validation Start")
    
    try:
        # 1. File checks
        if "file" not in frappe.request.files:
            return fail("NO_FILE", "No file uploaded")

        file = frappe.request.files["file"]

        if not file.filename.lower().endswith(".xlsx"):
            return fail("INVALID_FILE_TYPE", "Only .xlsx files are supported")

        file.seek(0, 2)
        size_mb = file.tell() / (1024 * 1024)
        file.seek(0)
        if size_mb > MAX_FILE_SIZE_MB:
            return fail("FILE_TOO_LARGE", f"File size exceeds {MAX_FILE_SIZE_MB} MB")

        print(f"[TIMING] File validation: {time.time() - start_time:.2f}s")

        # 2. Load workbook
        load_start = time.time()
        try:
            file_content = file.read()
            wb_input = load_workbook(BytesIO(file_content), data_only=True, read_only=False)
            print(f"[TIMING] Workbook loaded: {time.time() - load_start:.2f}s")
        except Exception as e:
            if "encrypted" in str(e).lower() or "password" in str(e).lower():
                return fail("PASSWORD_PROTECTED", "This file is password protected.")
            return fail("UNREADABLE_FILE", f"File is corrupted or unreadable: {str(e)}")

        # 3. Create output workbook
        wb_output = Workbook()
        wb_output.remove(wb_output.active)
        
        overall_stats = {
            "total_sheets": len(wb_input.sheetnames),
            "validated_sheets": 0,
            "total_errors": 0,
            "total_rows": 0,
            "sheet_results": [],
            "all_json_errors": []
        }
        
        print(f"\n[INFO] Processing {len(wb_input.sheetnames)} sheets: {wb_input.sheetnames}")
        
        # 4. Process each sheet
        for sheet_idx, sheet_name in enumerate(wb_input.sheetnames, 1):
            sheet_start = time.time()
            print(f"\n{'='*60}")
            print(f"[SHEET {sheet_idx}/{len(wb_input.sheetnames)}] {sheet_name}")
            print(f"{'='*60}")
            
            # Check timeout
            elapsed = time.time() - start_time
            if elapsed > VALIDATION_TIMEOUT:
                print(f"[WARNING] Overall timeout reached at {elapsed:.1f}s")
                break
            
            # Check DocType
            doctype_exists = check_doctype_exists(sheet_name)
            
            if not doctype_exists:
                print(f"[ERROR] DocType '{sheet_name}' not found")
                sheet_result = create_error_sheet(wb_output, sheet_name, "DOCTYPE_NOT_FOUND", 
                                                 f"DocType '{sheet_name}' not found in Frappe")
                overall_stats["sheet_results"].append(sheet_result)
                if sheet_result.get("json_errors"):
                    overall_stats["all_json_errors"].extend(sheet_result["json_errors"])
                continue
            
            print(f"[OK] DocType '{sheet_name}' exists")
            
            # Process sheet
            try:
                sheet_result = process_sheet_with_validation(
                    wb_input=wb_input,
                    wb_output=wb_output,
                    doctype=sheet_name,
                    skip_links=skip_links,
                    start_time=sheet_start
                )
                
                overall_stats["sheet_results"].append(sheet_result)
                
                if sheet_result.get("json_errors"):
                    overall_stats["all_json_errors"].extend(sheet_result["json_errors"])

                if sheet_result["success"]:
                    overall_stats["validated_sheets"] += 1
                
                # Add stats regardless of success status
                overall_stats["total_errors"] += sheet_result.get("error_count", 0)
                overall_stats["total_rows"] += sheet_result.get("total_rows", 0)
                
                sheet_time = time.time() - sheet_start
                print(f"\n[TIMING] Sheet '{sheet_name}' completed in {sheet_time:.2f}s")
                print(f"  - Rows processed: {sheet_result.get('total_rows', 0)}")
                print(f"  - Errors found: {sheet_result.get('error_count', 0)}")
                
            except ValidationTimeout:
                print(f"[ERROR] Sheet timeout after {SHEET_TIMEOUT}s")
                sheet_result = create_error_sheet(wb_output, sheet_name, "TIMEOUT_ERROR",
                                                 f"Sheet processing exceeded {SHEET_TIMEOUT}s timeout")
                overall_stats["sheet_results"].append(sheet_result)
            except Exception as e:
                print(f"[ERROR] Sheet processing failed: {str(e)}")
                traceback.print_exc()
                sheet_result = create_error_sheet(wb_output, sheet_name, "PROCESSING_ERROR",
                                                 f"Error: {str(e)}")
                overall_stats["sheet_results"].append(sheet_result)

        # 5. Save file
        print(f"\n{'='*60}")
        print("SAVING OUTPUT FILE")
        print(f"{'='*60}")
        
        save_start = time.time()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb_output.save(tmp.name)
            tmp.seek(0)
            content = tmp.read()
            tmp_path = tmp.name

        saved_file = save_file(
            fname=f"MultiSheet_Validated.xlsx",
            content=content,
            dt=None,
            dn=None,
            is_private=0
        )

        if os.path.exists(tmp_path):
            os.remove(tmp_path)
        
        total_time = time.time() - start_time
        
        summary = f"""
{'='*60}
VALIDATION COMPLETE
{'='*60}
File saved: {saved_file.file_url}
Total time: {total_time:.2f}s
Sheets processed: {overall_stats['validated_sheets']}/{overall_stats['total_sheets']}
Total rows: {overall_stats['total_rows']}
Total errors: {overall_stats['total_errors']}
{'='*60}
"""
        print(summary)
        frappe.log_error(summary, "Validation Complete")

        return safe_response({
            "structure_valid": overall_stats["total_errors"] == 0,
            "file_url": saved_file.file_url,
            "total_sheets": overall_stats["total_sheets"],
            "validated_sheets": overall_stats["validated_sheets"],
            "total_errors": overall_stats["total_errors"],
            "total_rows": overall_stats["total_rows"],
            "errors": overall_stats["all_json_errors"],
            "sheet_results": overall_stats["sheet_results"],
            "processing_time": round(total_time, 2)
        })
        
    except Exception as e:
        error_msg = f"[ERROR] Unexpected error: {str(e)}\n{traceback.format_exc()}"
        print(error_msg)
        frappe.log_error(error_msg, "Validation Error")
        return fail("PROCESSING_ERROR", f"Unexpected error: {str(e)}")


# ================= PROCESS SHEET =================

def process_sheet_with_validation(wb_input, wb_output, doctype, skip_links, start_time):
    """
    Process sheet with DETAILED link validation logging
    """
    json_errors = []

    try:
        # Get sheet
        sheet_map = {s.lower().replace(" ", ""): s for s in wb_input.sheetnames}
        sheet_key = doctype.lower().replace(" ", "")
        
        if sheet_key not in sheet_map:
            return create_error_sheet(wb_output, doctype, "SHEET_NOT_FOUND", 
                                     f"Worksheet '{doctype}' not found.")
        
        sheet_input = wb_input[sheet_map[sheet_key]]
        
        # Read headers
        header_row = next(sheet_input.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            return create_error_sheet(wb_output, doctype, "NO_HEADER", "Header row missing")

        headers = [str(h).strip() for h in header_row if h]
        if not headers:
            return create_error_sheet(wb_output, doctype, "EMPTY_HEADERS", 
                                     "The header row is empty or the sheet is blank.")

        print(f"[INFO] Found {len(headers)} columns: {headers}")

        # Validation checks
        if any(h is None or str(h).strip() == "" for h in header_row):
            return create_error_sheet(wb_output, doctype, "EMPTY_HEADERS", 
                                     "The header row contains blank cells between columns.")

        if len(headers) != len(set(headers)):
            return create_error_sheet(wb_output, doctype, "DUPLICATE_HEADERS", 
                                     "The header row contains duplicate column names.")

        # Create output sheet
        ws_output = wb_output.create_sheet(title=doctype)
        new_headers = ["Error Detected", "No. of Error", "DetailsMessage"] + headers
        ws_output.append(new_headers)
        
        for cell in ws_output[1]:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL

        # Load metadata
        print(f"[INFO] Loading DocType metadata...")
        meta, meta_err = safe_get_meta(doctype)
        if meta_err:
            return create_error_sheet(wb_output, doctype, "DOCTYPE_ERROR", 
                                     f"Failed to load DocType metadata")

        # LINK VALIDATION PREFETCH with detailed logging
        if not skip_links:
            print(f"\n[LINK VALIDATION] Starting prefetch...")
            prefetch_start = time.time()
            prefetch_link_caches(doctype, meta)
            prefetch_time = time.time() - prefetch_start
            print(f"[LINK VALIDATION] Prefetch completed in {prefetch_time:.2f}s\n")
        else:
            print(f"[INFO] Link validation SKIPPED\n")

        # Build field mappings
        headers_norm = {clean_header(h): h for h in headers}
        header_index_map = {h: i for i, h in enumerate(headers)}
        
        required_columns = []
        field_map = build_field_map(meta, headers)
        
        for df in meta.fields:
            if df.reqd and df.fieldtype not in ("Section Break", "Column Break", "Tab Break"):
                label_norm = clean_header(df.label) if df.label else ""
                fname_norm = clean_header(df.fieldname) if df.fieldname else ""
                if label_norm in headers_norm:
                    required_columns.append(headers_norm[label_norm])
                elif fname_norm in headers_norm:
                    required_columns.append(headers_norm[fname_norm])

        print(f"[INFO] Required columns: {required_columns}")

        unique_columns = get_unique_columns(doctype, headers, meta)
        primary_key = get_primary_key(doctype, headers, meta)

        seen_rows = set()
        seen_unique = {c: set() for c in unique_columns}
        seen_primary = set()
        
        total_errors = 0

        # Load all rows
        all_rows = list(sheet_input.iter_rows(min_row=2, values_only=True))
        
        # Check if rows are actually empty (sometimes iter_rows returns empty tuples)
        valid_rows = [r for r in all_rows if any(c is not None and str(c).strip() != "" for c in r)]
        
        if not valid_rows:
             return create_error_sheet(wb_output, doctype, "NO_DATA_ROWS", 
                                      "The file contains no data rows.")
                                      
        print(f"[INFO] Processing {len(all_rows)} data rows...")

        # Process rows
        validation_start = time.time()
        for row_idx, row in enumerate(all_rows, start=2):
            # Timeout check every 100 rows
            if row_idx % 100 == 0:
                elapsed = time.time() - start_time
                if elapsed > SHEET_TIMEOUT:
                    print(f"[TIMEOUT] Sheet timeout at row {row_idx}")
                    raise ValidationTimeout(f"Sheet {doctype} timeout")
                print(f"  [PROGRESS] Processed {row_idx-1} rows ({elapsed:.1f}s elapsed)")
            
            row_dict = dict(zip(headers, row))
            
            short_errors = []
            detailed_errors = []
            failed_columns = set()

            # Empty row check
            if all(v is None or str(v).strip() in ("", "NA", "N/A", "na", "n/a") for v in row_dict.values()):
                err_code = "EMPTY_ROW"
                msg = "Row is completely empty"
                short_errors.append("Empty row")
                detailed_errors.append(msg)
                
                json_errors.append({
                    "sheet": doctype,
                    "row": row_idx,
                    "column": "Entire Row",
                    "code": "Empty Row",
                    "message": msg,
                    "value_entered": "",
                    "error_type": "Empty Row"
                })
            else:
                # Required fields
                for col in required_columns:
                    val = row_dict.get(col)
                    if val in (None, "", "NA", "N/A"):
                        err_code = "REQUIRED_FIELD_EMPTY"
                        msg = f"Required field '{col}' is empty"
                        
                        short_errors.append(f"{col}: Required field empty")
                        detailed_errors.append(msg)
                        failed_columns.add(col)

                        json_errors.append({
                            "sheet": doctype,
                            "row": row_idx,
                            "column": col,
                            "code": get_readable_error_code(err_code),
                            "message": msg,
                            "value_entered": "",
                            "error_type": get_formatted_error_type(err_code)
                        })

                # Datatype & Link Validation (with logging)
                dtype_errors = validate_datatypes(doctype, row_dict, row_idx, headers, meta, skip_links, field_map)
                if dtype_errors:
                    for err in dtype_errors:
                        col = err.get("column", "")
                        code = err.get("code", "")
                        raw_msg = err.get("message", "")
                        
                        readable_code = get_readable_error_code(code).strip()
                        
                        if col:
                            short_errors.append(f"{col}: {readable_code}")
                            failed_columns.add(col)
                        else:
                            short_errors.append(readable_code)

                        detail_msg = raw_msg
                        suggestions = err.get("suggestions", [])
                        if suggestions:
                            detail_msg += f" Suggest: {', '.join(str(s) for s in suggestions)}"
                        
                        if col:
                            detail_msg = f"{col}: {detail_msg}"
                        
                        detailed_errors.append(detail_msg)

                        val_entered = row_dict.get(col) if col else ""
                        err_object = {
                            "sheet": doctype,
                            "row": row_idx,
                            "column": col,
                            "code": get_readable_error_code(code),
                            "message": raw_msg,
                            "value_entered": str(val_entered) if val_entered is not None else "",
                            "error_type": get_formatted_error_type(code)
                        }
                        if suggestions:
                            err_object["suggestions"] = suggestions
                        json_errors.append(err_object)

                # Duplicate row check
                row_values = tuple((str(v) if v not in (None, "", "NA", "N/A") else "") for v in row_dict.values())
                if row_values and row_values in seen_rows:
                    short_errors.append("Duplicate row")
                    detailed_errors.append("Duplicate row")
                    
                    json_errors.append({
                        "sheet": doctype,
                        "row": row_idx,
                        "column": "Entire Row",
                        "code": get_readable_error_code("DUPLICATE_ROW"),
                        "message": "This row is a duplicate of a previous row",
                        "value_entered": "Row Data",
                        "error_type": get_formatted_error_type("DUPLICATE_ROW")
                    })
                seen_rows.add(row_values)

                # Primary key duplicate
                if primary_key:
                    pk_val = row_dict.get(primary_key)
                    if pk_val and pk_val not in (None, "", "NA", "N/A"):
                        pk_str = str(pk_val).strip()
                        if pk_str in seen_primary:
                            err_code = "DUPLICATE_PRIMARY_KEY"
                            msg = f"{primary_key}: Duplicate ID ({pk_val})"
                            
                            short_errors.append(f"{primary_key}: Duplicate ID")
                            detailed_errors.append(msg)
                            failed_columns.add(primary_key)

                            json_errors.append({
                                "sheet": doctype,
                                "row": row_idx,
                                "column": primary_key,
                                "code": get_readable_error_code(err_code),
                                "message": msg,
                                "value_entered": str(pk_val),
                                "error_type": get_formatted_error_type(err_code)
                            })
                        seen_primary.add(pk_str)

                # Unique value duplicate
                for col in unique_columns:
                    val = row_dict.get(col)
                    if val and val not in (None, "", "NA", "N/A"):
                        val_str = str(val).strip()
                        if val_str in seen_unique[col]:
                            err_code = "DUPLICATE_UNIQUE"
                            msg = f"{col}: Duplicate value ({val})"
                            
                            short_errors.append(f"{col}: Duplicate value")
                            detailed_errors.append(msg)
                            failed_columns.add(col)
                            
                            json_errors.append({
                                "sheet": doctype,
                                "row": row_idx,
                                "column": col,
                                "code": get_readable_error_code(err_code),
                                "message": msg,
                                "value_entered": str(val),
                                "error_type": get_formatted_error_type(err_code)
                            })

                        seen_unique[col].add(val_str)

            # Write row
            has_error = len(short_errors) > 0
            error_count = len(short_errors)
            
            if has_error:
                total_errors += 1
            
            error_val = "Error" if has_error else "No Error"
            detail_val = "; ".join(detailed_errors) if detailed_errors else "No Error"
            
            output_row_data = [error_val, error_count, detail_val]
            for h in headers:
                val = row_dict.get(h)
                output_row_data.append(val)
            
            ws_output.append(output_row_data)
            
            # Highlighting
            current_excel_row = ws_output.max_row
            
            if has_error:
                cell = ws_output.cell(row=current_excel_row, column=1)
                cell.font = ERROR_TEXT_COLOR

                cell_count = ws_output.cell(row=current_excel_row, column=2)
                cell_count.font = ERROR_TEXT_COLOR
            
            for col_name in failed_columns:
                if col_name in header_index_map:
                    col_idx = header_index_map[col_name]
                    excel_col_idx = col_idx + 4
                    
                    cell = ws_output.cell(row=current_excel_row, column=excel_col_idx)
                    cell.fill = ERROR_FILL
                    cell.font = ERROR_TEXT_COLOR

        validation_time = time.time() - validation_start
        print(f"[TIMING] Row validation completed in {validation_time:.2f}s")

        return {
            "sheet_name": doctype,
            "success": True,
            "error_count": len(json_errors),
            "total_rows": len(all_rows),
            "json_errors": json_errors
        }
        
    except ValidationTimeout:
        raise
    except Exception as e:
        error_details = traceback.format_exc()
        print(f"[ERROR] {error_details}")
        return create_error_sheet(wb_output, doctype, "PROCESSING_ERROR", 
                                 f"Unexpected error: {str(e)}")


# ================= LINK VALIDATION WITH LOGGING =================

def prefetch_link_caches(doctype, meta=None):
    """
    WITH DETAILED LOGGING: Shows exactly what's happening
    """
    if meta is None:
        try:
            meta = frappe.get_meta(doctype)
        except:
            return
    
    link_doctypes = set()
    for df in meta.fields:
        if df.fieldtype == "Link" and df.options:
            link_doctypes.add(df.options)
    
    if not link_doctypes:
        print(f"  [LINK] No link fields found")
        return
    
    print(f"  [LINK] Found {len(link_doctypes)} link fields to process")
    
    for idx, linked_dt in enumerate(link_doctypes, 1):
        try:
            print(f"  [{idx}/{len(link_doctypes)}] {linked_dt}:")
            
            # Count
            count_start = time.time()
            count = frappe.db.count(linked_dt)
            count_time = time.time() - count_start
            print(f"      Count: {count:,} records ({count_time:.2f}s)")
            
            _link_cache_sizes[linked_dt] = count
            
            if count > LINK_CACHE_LIMIT:
                print(f"      ⚠️  SKIPPING cache (exceeds {LINK_CACHE_LIMIT:,} limit)")
                print(f"      Will use direct DB queries")
                continue
            
            # Cache
            cache_start = time.time()
            _link_cache[linked_dt] = set(frappe.get_all(linked_dt, pluck="name"))
            cache_time = time.time() - cache_start
            
            print(f"      ✅ Cached {len(_link_cache[linked_dt]):,} records ({cache_time:.2f}s)")
            
        except Exception as e:
            print(f"      ❌ ERROR: {str(e)}")


def link_exists(doctype, name, case_sensitive=False):
    """
    WITH LOGGING for large table queries
    """
    if not doctype or not name:
        return False
    
    try:
        if not frappe.db.exists("DocType", doctype):
            return False
        
        # Large table check
        is_large = doctype in _link_cache_sizes and _link_cache_sizes[doctype] > LINK_CACHE_LIMIT
        
        if is_large:
            # Direct query (logged only first time per DocType)
            name_str = str(name).strip()
            result = frappe.db.exists(doctype, name_str)
            return bool(result)
        
        # Cache-based lookup
        if doctype not in _link_cache:
            try:
                count = frappe.db.count(doctype)
                if count > LINK_CACHE_LIMIT:
                    _link_cache_sizes[doctype] = count
                    result = frappe.db.exists(doctype, name)
                    return bool(result)
                else:
                    _link_cache[doctype] = set(frappe.get_all(doctype, pluck="name"))
            except:
                return False
        
        name_str = str(name).strip()
        if case_sensitive:
            return name_str in _link_cache[doctype]
        
        name_lower = name_str.lower()
        for cached_name in _link_cache[doctype]:
            if str(cached_name).lower() == name_lower:
                return True
        return False
    except Exception:
        return False


def get_link_suggestions(doctype, name, max_suggestions=3):
    if doctype not in _link_cache:
        try:
            link_exists(doctype, "__init__")
        except:
            return []
    name_lower = str(name).lower()
    suggestions = []
    if doctype in _link_cache:
        for cached_name in _link_cache[doctype]:
            if name_lower in str(cached_name).lower() or str(cached_name).lower() in name_lower:
                suggestions.append(cached_name)
                if len(suggestions) >= max_suggestions:
                    break
    return suggestions


# ================= UTILITY FUNCTIONS (same as before) =================

def create_error_sheet(wb_output, sheet_name, error_code, error_message):
    ws_output = wb_output.create_sheet(title=sheet_name)
    ws_output.append(["ERROR", error_message])
    ws_output["A1"].fill = ERROR_FILL
    ws_output["A1"].font = ERROR_TEXT_COLOR
    ws_output["B1"].fill = ERROR_FILL
    ws_output["B1"].font = ERROR_TEXT_COLOR
    
    json_error = {
        "sheet": sheet_name,
        "row": 0,
        "column": "Sheet",
        "code": get_readable_error_code(error_code),
        "message": error_message,
        "value_entered": "",
        "error_type": get_formatted_error_type(error_code)
    }

    return {
        "sheet_name": sheet_name,
        "success": False,
        "error": error_code,
        "message": error_message,
        "error_count": 1,
        "total_rows": 0,
        "json_errors": [json_error]
    }


def check_doctype_exists(doctype_name):
    try:
        return frappe.db.exists("DocType", doctype_name)
    except Exception:
        return False


def get_readable_error_code(code):
    return ERROR_CODE_LABELS.get(code, code.replace('_', ' '))


def get_formatted_error_type(code):
    return ERROR_CODE_LABELS.get(code, code.replace('_', ' ').title())


def clean_header(h):
    if h is None:
        return ""
    base = re.sub(r"\s*\(.*\)$", "", str(h).strip(), flags=re.IGNORECASE)
    return base.lower().replace(" ", "")


def safe_response(obj):
    try:
        return json.loads(json.dumps(obj, default=str, ensure_ascii=False))
    except Exception:
        return {"structure_valid": False, "errors": [{"code": "RESPONSE_ERROR", "message": "Serialization Error"}]}


def fail(code, message):
    err_type = ERROR_CODE_LABELS.get(code, code)
    return safe_response({
        "structure_valid": False, 
        "errors": [{
            "code": code, 
            "message": message,
            "error_type": err_type
        }]
    })



def convert_to_frappe_format(validation_result):
    message = {}
    
    for sheet_result in validation_result.get("sheet_results", []):
        sheet_name = sheet_result.get("sheet_name")
        errors = sheet_result.get("json_errors", [])
        
        logs = []
        warnings = []
        
        for err in errors:
            row = err.get("row", 0)
            
            if row > 0:
                # Regular error log
                logs.append({
                    "rows": [row],
                    "status": "error",
                    "document": None,
                    "message": err.get("message", "")
                })
            else:
                # Column-level warning
                col = err.get("column", "")
                col_num = int(col.replace("Column ", "")) if "Column" in col else 0
                
                warnings.append({
                    "col": col_num,
                    "message": err.get("message", ""),
                    "type": "warning"
                })
        
        message[sheet_name] = {
            "data_import": f"{sheet_name} Import on {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')}",
            "status": "Error" if errors else "Success",
            "logs": logs,
            "warnings": warnings
        }
    
    return {"message": message}


def safe_get_meta(doctype):
    try:
        return frappe.get_meta(doctype), None
    except Exception as e:
        return None, fail("DOCTYPE_ERROR", str(e))


def get_unique_columns(doctype, headers, meta):
    headers_norm = {clean_header(h): h for h in headers}
    cols = []
    for df in meta.fields:
        if getattr(df, "unique", False):
            label_norm = clean_header(df.label) if df.label else ""
            fname_norm = clean_header(df.fieldname) if df.fieldname else ""
            if label_norm in headers_norm:
                cols.append(headers_norm[label_norm])
            elif fname_norm in headers_norm:
                cols.append(headers_norm[fname_norm])
    return cols


def get_primary_key(doctype, headers, meta):
    headers_norm = {clean_header(h): h for h in headers}
    for df in meta.fields:
        if df.fieldname == "name":
            label_norm = clean_header(df.label) if df.label else ""
            if label_norm in headers_norm:
                return headers_norm[label_norm]
            if "name" in headers_norm:
                return headers_norm["name"]
    return None


def build_field_map(meta, headers=None):
    print(f"[DEBUG] build_field_map called. Meta Fields: {len(meta.fields)} Headers: {len(headers) if headers else 0}")
    field_map = {}
    
    # Use meta.fields if available
    for df in meta.fields:
        if getattr(df, "label", None):
            k = clean_header(df.label)
            field_map[k] = df
            # print(f"[DEBUG-MAP] Added key: '{k}' from Label: '{df.label}'")
        if getattr(df, "fieldname", None):
            k = clean_header(df.fieldname)
            if k not in field_map:
                field_map[k] = df
    
    # If meta.fields is empty and headers provided, create fields from headers
    if not field_map and headers:
        for h in headers:
            key = clean_header(h)
            h_lower = str(h).lower()
            
            # Determine fieldtype based on header name
            if key == "year":
                fieldtype = "Int"
            elif "date" in h_lower:
                fieldtype = "Date"
            # Prioritize Financial/Amount columns
            elif re.search(r'\b(financial|budget|amount|cost|price|rate)\b', h_lower):
                fieldtype = "Currency" if "rate" in h_lower or "price" in h_lower or "financial" in h_lower else "Float"
                print(f"[DEBUG-MAP] Header: {h} -> Type: {fieldtype}")
            
            # Other numeric indicators
            elif re.search(r'\b(count|total|value|qty|quantity|ranking|index|strength|position|vacant|ratio|currency)\b', h_lower):
                # Skip if it's a text column like "Case Number & Title"
                if "&" in h_lower or "title" in h_lower or "case" in h_lower or "name" in h_lower:
                    fieldtype = "Data"
                else:
                    fieldtype = "Float"
            else:
                fieldtype = "Data"
            
            # Create simple field object
            class SimpleField:
                pass
            df = SimpleField()
            df.label = h
            df.fieldname = key
            df.fieldtype = fieldtype
            df.reqd = 0
            df.unique = 0
            df.options = None
            
            field_map[key] = df
    
    return field_map



def validate_datatypes(doctype, row_dict, row_idx, headers, meta, skip_link_validation, field_map=None):
    if field_map is None:
        field_map = build_field_map(meta, headers)

    errors = []
    
    # Columns that can be empty (system fields, notes, descriptions)
    optional_columns = ("id", "institute", "notes", "description", "brief", "details", "remarks", "comments")
    
    for label, value in row_dict.items():
        original_value = value
        key = clean_header(label)
        
        # Skip checking system/optional columns for emptiness
        is_optional = key in optional_columns or label.lower().startswith("id ")
        
        # Check for empty values in non-optional columns
        # Treat whitespace as empty
        str_val = str(value).strip() if value is not None else ""
        
        if str_val == "" or str_val in ("NA", "N/A", "na", "n/a"):
            if not is_optional:
                errors.append({
                    "row": row_idx, 
                    "column": label, 
                    "code": "REQUIRED_FIELD_EMPTY", 
                    "message": f"Field '{label}' is empty"
                })
            continue
        
        df = field_map.get(key)
        
        # Fallback for Year column if not in map but detected by name
        is_year_by_name = "year" in label.lower() and len(label) < 15
        
        if not df:
            if is_year_by_name:
                class DummyField: fieldtype = "Int"; fieldname = "year"
                df = DummyField()
            else:
                continue
        
        ft = df.fieldtype
        
        if "financial" in label.lower() and ft == "Currency":
             pass

        # Year validation
        is_year_field = key == "year" or key.endswith("_year") or getattr(df, "fieldname", "") == "year" or is_year_by_name

        if is_year_field:
            y = validate_year_value(value)
            if y is None:
                errors.append({"row": row_idx, "column": label, "code": "INVALID_YEAR", "message": f"Invalid year: {value}"})
                continue
            if not (1900 <= int(y) <= datetime.date.today().year + 1):
                # print(f"[DEBUG-YEAR-CHECK] FAIL RANGE: {y}")
                errors.append({"row": row_idx, "column": label, "code": "INVALID_YEAR_RANGE", "message": f"Year {y} out of allowed range 1900-{datetime.date.today().year + 1}"})
            if ft != "Link":
                continue

        # Link validation
        if ft == "Link":
            link_err = validate_link_field(df, value, label, row_idx, skip_link_validation, original_value)
            if link_err:
                errors.append(link_err)
            continue
        
        # Numeric/Date checks
        if ft in ("Int", "Check"):
            # Try to validate/convert to integer
            is_valid_int = False
            if isinstance(value, (int, bool)):
                is_valid_int = True
            elif isinstance(value, float) and value.is_integer():
                is_valid_int = True
            elif isinstance(value, str):
                try:
                    # Try to convert string to number
                    num_val = float(value.strip())
                    if num_val.is_integer():
                        is_valid_int = True
                except (ValueError, AttributeError):
                    pass
            
            if not is_valid_int:
                errors.append({"row": row_idx, "column": label, "code": "INVALID_INT", "message": f"'{value}' is not a valid number"})
        
        elif ft in ("Float", "Currency", "Percent"):
            # Try to validate/convert to float
            is_valid_float = False
            if isinstance(value, (int, float)):
                is_valid_float = True
            elif isinstance(value, str):
                try:
                    # Try to convert string to float
                    float(value.strip())
                    is_valid_float = True
                except (ValueError, AttributeError):
                    pass
            
            if not is_valid_float:
                errors.append({"row": row_idx, "column": label, "code": "INVALID_FLOAT", "message": f"'{value}' is not a valid number"})
        elif ft == "Date":
            if not isinstance(value, (datetime.date, datetime.datetime)):
                errors.append({"row": row_idx, "column": label, "code": "INVALID_DATE", "message": "Invalid date format"})
        elif ft == "Datetime":
            if not isinstance(value, datetime.datetime):
                errors.append({"row": row_idx, "column": label, "code": "INVALID_DATETIME", "message": "Invalid datetime format"})

    return errors


def validate_year_value(value):
    if isinstance(value, (datetime.date, datetime.datetime)):
        return value.year
    if isinstance(value, (int, float)):
        return int(value)
    if isinstance(value, str) and re.fullmatch(r"\d{4}", value.strip()):
        return int(value.strip())
    try:
        return datetime.datetime.fromisoformat(str(value)).year
    except:
        return None


def validate_link_field(df, value, label, row_idx, skip_link_validation, original_value):
    if skip_link_validation or not value:
        return None
    
    linked_doctype = df.options
    if not linked_doctype:
        return {
            "row": row_idx,
            "column": label,
            "code": "LINK_CONFIG_ERROR",
            "message": "No target DocType"
        }

    value_str = str(value).strip()

    # 1️⃣ Direct DB exists check first (most reliable)
    if frappe.db.exists(linked_doctype, value_str):
        return None

    # 2️⃣ Case-insensitive cache check (if cache exists)
    existing = _link_cache.get(linked_doctype)
    if existing:
        for v in existing:
            if str(v).strip().lower() == value_str.lower():
                return None

    # 3️⃣ Suggestions (optional)
    suggestions = get_link_suggestions(linked_doctype, value_str)

    return {
        "row": row_idx,
        "column": label,
        "code": "LINK_NOT_FOUND",
        "message": f"{label}: '{original_value}' not found in {linked_doctype}",
        "suggestions": suggestions
    }
